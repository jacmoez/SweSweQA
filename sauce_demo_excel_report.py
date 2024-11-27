from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import time


def report_excels(file_name="report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"
    ws.append(["Test", "Message"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4dca35", end_color="4dca35", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(file_name)

def write_report_excel(test="", msg="", file_name="report.xlsx"):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.append([test,msg])
    wb.save(file_name)

class SwagTst:
    driver = webdriver.Firefox()

    def test_open_browser(self):
        self.driver.get("https://www.saucedemo.com/")
        write_report_excel("Test 1 :"," Open browser success!")

    def test_login_test(self):
        self.driver.find_element(By.ID, "user-name").send_keys("standard_user")
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//*[@id='password']").send_keys("secret_sauce", Keys.RETURN)
        write_report_excel("Test 2 : ","Login Success!")

    def test_check_page_go(self):
        expected_url = "https://www.saucedemo.com/inventory.html"
        actual_url = self.driver.current_url

        assert expected_url == actual_url
        assert "Swag Labs" in self.driver.title
        write_report_excel("Test 3 :", " verified success")

    def test_add_to_cart(self):
        self.driver.find_element(By.ID, "add-to-cart-sauce-labs-backpack").click()
        time.sleep(2)
        self.driver.find_element(By.ID, "add-to-cart-sauce-labs-bolt-t-shirt").click()
        time.sleep(2)
        self.driver.find_element(By.ID, "add-to-cart-sauce-labs-fleece-jacket").click()
        write_report_excel("Test 4 :"," Add item to Cart")

    def test_remove_item(self):
        self.driver.find_element(By.ID, "remove-sauce-labs-fleece-jacket").click()
        write_report_excel("Test 5 :" ,"Remove 1 item")

    def test_view_order_items(self):
        self.driver.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        write_report_excel("Test 6 :"," View Order Item")

    def test_check_cart_page_go(self):
        expected_url = "https://www.saucedemo.com/cart.html"
        actual_url = self.driver.current_url

        assert expected_url == actual_url
        assert "Swag Labs" in self.driver.title
        write_report_excel("Test 7 :"," check cart page success")

    def test_view_cart_ditail(self):
        self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div/div[1]/div[3]/div[2]/a/div").click()

        self.driver.find_element(By.CLASS_NAME, "shopping_cart_link").click()

        self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div/div[1]/div[4]/div[2]/a/div").click()

        self.driver.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        write_report_excel("Test 8 :"," view cart ditail Success!")

    def test_check_out(self):
        self.driver.find_element(By.ID, "checkout").click()
        self.driver.find_element(By.ID, "first-name").send_keys("QA")
        time.sleep(2)
        self.driver.find_element(By.ID, "last-name").send_keys("Swe Swe")
        time.sleep(2)
        self.driver.find_element(By.ID, "postal-code").send_keys("11111")
        time.sleep(2)
        self.driver.find_element(By.ID, "continue").click()
        write_report_excel("Test 9 :"," Check Out Success!")

    def test_check_checkout(self):
        item_total = self.driver.find_element(By.CLASS_NAME, "summary_subtotal_label").text
        tax = self.driver.find_element(By.CLASS_NAME, "summary_tax_label").text
        total = self.driver.find_element(By.CLASS_NAME, "summary_total_label").text
        write_report_excel(f"{item_total}.")
        write_report_excel(f"{tax}.")
        write_report_excel(f"{total}")
        #write_report_excel("="*30)

    def finish(self):
        self.driver.find_element(By.ID, "finish").click()
        complete = self.driver.find_element(By.CLASS_NAME, "complete-header").text
        time.sleep(2)
        write_report_excel(complete)
        self.driver.find_element(By.ID, "back-to-products").click()
        write_report_excel('Test 10 :"," Finish Success')
        #write_report_excel("="*50)

    def main(self):
        report_excels()
        self.test_open_browser()
        self.test_login_test()
        self.test_check_page_go()
        self.test_add_to_cart()
        self.test_remove_item()
        self.test_view_order_items()
        self.test_check_cart_page_go()
        self.test_view_cart_ditail()
        self.test_check_out()
        self.test_check_checkout()
        self.finish()


SwagTst().main()
